import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import csv
import os

try:
    import openpyxl
except ImportError:
    openpyxl = None

CONDITIONS = [
    ("", ""),  # No filter
    ("==", "Equals"),
    (">", "Greater than"),
    ("<", "Less than"),
    (">=", "Greater or equal"),
    ("<=", "Less or equal"),
    ("contains", "Contains"),
    ("not contains", "Not contains"),
]

class FilterTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.file_var = tk.StringVar()
        self.limit_var = tk.StringVar(value="100")
        self.all_columns = []
        self.selected_columns = []
        self.filter_widgets = []
        self.count_var = tk.StringVar(value="Matching rows: 0")
        self.setup_ui()

    def setup_ui(self):
        tk.Label(self, text="CSV/Excel File:").grid(row=0, column=0, sticky="e")
        tk.Entry(self, textvariable=self.file_var, width=40).grid(row=0, column=1)
        tk.Button(self, text="Browse...", command=self.select_file).grid(row=0, column=2)

        tk.Label(self, text="Max Rows (leave blank for all):").grid(row=1, column=0, sticky="e")
        limit_entry = tk.Entry(self, textvariable=self.limit_var, width=15)
        limit_entry.grid(row=1, column=1, sticky="w")
        limit_entry.bind("<KeyRelease>", lambda e: self.update_count_label())

        tk.Label(self, text="Select Column:").grid(row=2, column=0, sticky="e")
        self.column_select = ttk.Combobox(self, values=self.all_columns, state="readonly", width=30)
        self.column_select.grid(row=2, column=1, sticky="w")
        tk.Button(self, text="Add", command=self.add_column).grid(row=2, column=2, sticky="w")

        self.selected_columns_frame = tk.Frame(self)
        self.selected_columns_frame.grid(row=3, column=0, columnspan=3, pady=10)

        self.count_label = tk.Label(self, textvariable=self.count_var, fg="blue")
        self.count_label.grid(row=5, column=0, columnspan=3, pady=5)

        tk.Button(self, text="Process", command=self.process_file).grid(row=4, column=1, pady=10)

    def select_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("CSV or Excel files", "*.csv *.xlsx")],
            title="Select a CSV or Excel file"
        )
        if file_path:
            ext = os.path.splitext(file_path)[1].lower()
            if ext not in [".csv", ".xlsx"]:
                messagebox.showerror("Error", "Please select a CSV (.csv) or Excel (.xlsx) file.")
                return
            if ext == ".xlsx" and openpyxl is None:
                messagebox.showerror("Error", "openpyxl is required for Excel files. Please install it with 'pip install openpyxl'.")
                return
            self.file_var.set(file_path)
            self.load_headers(file_path)
            self.reset_column_selection()
            self.update_count_label()

    def load_headers(self, file_path):
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".csv":
            with open(file_path, 'r', newline='', encoding='utf-8') as infile:
                reader = csv.reader(infile)
                headers = next(reader)
        elif ext == ".xlsx":
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        else:
            headers = []
        self.all_columns.clear()
        self.all_columns.extend(headers)
        self.column_select['values'] = self.all_columns
        self.column_select.set('')
        self.reset_column_selection()

    def reset_column_selection(self):
        for widget in self.selected_columns_frame.winfo_children():
            widget.destroy()
        self.selected_columns.clear()
        self.filter_widgets.clear()
        self.update_count_label()

    def add_column(self):
        col = self.column_select.get()
        if not col or col in self.selected_columns:
            return
        self.selected_columns.append(col)
        idx = len(self.selected_columns) - 1

        row = idx
        tk.Label(self.selected_columns_frame, text=col).grid(row=row, column=0, padx=2, pady=2)
        cond_var = tk.StringVar(value="")
        cond_menu = ttk.Combobox(self.selected_columns_frame, textvariable=cond_var, width=12, state="readonly")
        cond_menu['values'] = [label for code, label in CONDITIONS]
        cond_menu.current(0)
        cond_menu.grid(row=row, column=1, padx=2, pady=2)
        entry = tk.Entry(self.selected_columns_frame, width=14)
        entry.grid(row=row, column=2, padx=2, pady=2)
        remove_btn = tk.Button(self.selected_columns_frame, text="Remove", command=lambda: self.remove_column(col))
        remove_btn.grid(row=row, column=3, padx=2, pady=2)
        self.filter_widgets.append((col, cond_var, entry))

        cond_menu.bind("<<ComboboxSelected>>", lambda e: self.update_count_label())
        entry.bind("<KeyRelease>", lambda e: self.update_count_label())
        self.update_count_label()

    def remove_column(self, col):
        if col in self.selected_columns:
            idx = self.selected_columns.index(col)
            self.selected_columns.pop(idx)
            self.filter_widgets.pop(idx)
            for widget in self.selected_columns_frame.winfo_children():
                widget.destroy()
            for i, (col, cond_var, entry) in enumerate(self.filter_widgets):
                tk.Label(self.selected_columns_frame, text=col).grid(row=i, column=0, padx=2, pady=2)
                cond_menu = ttk.Combobox(self.selected_columns_frame, textvariable=cond_var, width=12, state="readonly")
                cond_menu['values'] = [label for code, label in CONDITIONS]
                cond_menu.current(0)
                cond_menu.grid(row=i, column=1, padx=2, pady=2)
                entry.grid(row=i, column=2, padx=2, pady=2)
                remove_btn = tk.Button(self.selected_columns_frame, text="Remove", command=lambda c=col: self.remove_column(c))
                remove_btn.grid(row=i, column=3, padx=2, pady=2)
                cond_menu.bind("<<ComboboxSelected>>", lambda e: self.update_count_label())
                entry.bind("<KeyRelease>", lambda e: self.update_count_label())
            self.update_count_label()

    def row_matches_filters(self, row, filters, header):
        for col, cond_var, entry in filters:
            if col not in header:
                return False
            idx = header.index(col)
            value = row[idx]
            cond_label = cond_var.get()
            cond_code = ""
            for code, label in CONDITIONS:
                if label == cond_label:
                    cond_code = code
                    break
            filter_value = entry.get().strip()
            if not cond_code or not filter_value:
                continue
            try:
                if cond_code == "==":
                    if str(value) != filter_value:
                        return False
                elif cond_code == ">":
                    if float(value) <= float(filter_value):
                        return False
                elif cond_code == "<":
                    if float(value) >= float(filter_value):
                        return False
                elif cond_code == ">=":
                    if float(value) < float(filter_value):
                        return False
                elif cond_code == "<=":
                    if float(value) > float(filter_value):
                        return False
                elif cond_code == "contains":
                    if filter_value.lower() not in str(value).lower():
                        return False
                elif cond_code == "not contains":
                    if filter_value.lower() in str(value).lower():
                        return False
            except Exception:
                return False
        return True

    def get_filtered_rows(self, input_file, filters, max_records=None):
        ext = os.path.splitext(input_file)[1].lower()
        if ext == ".csv":
            with open(input_file, 'r', newline='', encoding='utf-8') as infile:
                reader = csv.reader(infile)
                header = next(reader)
                filtered_rows = (row for row in reader if self.row_matches_filters(row, filters, header))
                if max_records is not None:
                    limited_rows = [row for _, row in zip(range(max_records), filtered_rows)]
                else:
                    limited_rows = list(filtered_rows)
        elif ext == ".xlsx":
            wb = openpyxl.load_workbook(input_file, read_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = next(rows)
            filtered_rows = (row for row in rows if self.row_matches_filters(row, filters, header))
            if max_records is not None:
                limited_rows = [row for _, row in zip(range(max_records), filtered_rows)]
            else:
                limited_rows = list(filtered_rows)
        else:
            header = []
            limited_rows = []
        return header, limited_rows

    def update_count_label(self, *args):
        input_file = self.file_var.get()
        limit = self.limit_var.get().strip()
        ext = os.path.splitext(input_file)[1].lower()
        if not input_file or not os.path.isfile(input_file) or ext not in [".csv", ".xlsx"]:
            self.count_var.set("Matching rows: 0")
            return
        if limit:
            try:
                max_records = int(limit)
                if max_records <= 0:
                    raise ValueError
            except ValueError:
                self.count_var.set("Matching rows: 0")
                return
        else:
            max_records = None

        filters = self.filter_widgets
        try:
            header, filtered_rows = self.get_filtered_rows(input_file, filters, max_records)
            row_count = len(filtered_rows)
        except Exception:
            row_count = 0
        self.count_var.set(f"Matching rows: {row_count}")

    def process_file(self):
        input_file = self.file_var.get()
        limit = self.limit_var.get().strip()
        ext = os.path.splitext(input_file)[1].lower()
        if not input_file or not os.path.isfile(input_file) or ext not in [".csv", ".xlsx"]:
            messagebox.showerror("Error", "Please select a valid CSV (.csv) or Excel (.xlsx) file.")
            return
        if ext == ".xlsx" and openpyxl is None:
            messagebox.showerror("Error", "openpyxl is required for Excel files. Please install it with 'pip install openpyxl'.")
            return
        if limit:
            try:
                max_records = int(limit)
                if max_records <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid positive integer for the row limit, or leave blank for all.")
                return
        else:
            max_records = None

        filters = self.filter_widgets

        try:
            header, filtered_rows = self.get_filtered_rows(input_file, filters, max_records)
            row_count = len(filtered_rows)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while filtering:\n{e}")
            return

        if row_count == 0:
            messagebox.showinfo("No Rows", "No rows match the given filters.")
            return

        confirm = messagebox.askyesno(
            "Confirm Save",
            f"{row_count} row(s) match your filters and will be saved.\n\nDo you want to continue?"
        )
        if not confirm:
            return

        if ext == ".csv":
            filetypes = [("CSV files", "*.csv")]
            defaultext = ".csv"
            save_func = save_csv
        else:
            filetypes = [("Excel files", "*.xlsx")]
            defaultext = ".xlsx"
            save_func = save_xlsx

        output_file = filedialog.asksaveasfilename(
            defaultextension=defaultext,
            filetypes=filetypes,
            title="Save filtered file as"
        )
        if not output_file:
            return

        try:
            save_func(header, filtered_rows, output_file)
            messagebox.showinfo("Success", f"Filtered file saved to:\n{output_file}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")

class UpdateTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.file_var = tk.StringVar()
        self.limit_var = tk.StringVar(value="100")
        self.all_columns = []
        self.selected_columns = []
        self.update_widgets = []
        self.count_var = tk.StringVar(value="Matching rows: 0")
        self.setup_ui()

    def setup_ui(self):
        tk.Label(self, text="CSV/Excel File:").grid(row=0, column=0, sticky="e")
        tk.Entry(self, textvariable=self.file_var, width=40).grid(row=0, column=1)
        tk.Button(self, text="Browse...", command=self.select_file).grid(row=0, column=2)

        tk.Label(self, text="Max Rows (leave blank for all):").grid(row=1, column=0, sticky="e")
        limit_entry = tk.Entry(self, textvariable=self.limit_var, width=15)
        limit_entry.grid(row=1, column=1, sticky="w")
        limit_entry.bind("<KeyRelease>", lambda e: self.update_count_label())

        tk.Label(self, text="Select Column:").grid(row=2, column=0, sticky="e")
        self.column_select = ttk.Combobox(self, values=self.all_columns, state="readonly", width=30)
        self.column_select.grid(row=2, column=1, sticky="w")
        tk.Button(self, text="Add", command=self.add_column).grid(row=2, column=2, sticky="w")

        self.selected_columns_frame = tk.Frame(self)
        self.selected_columns_frame.grid(row=3, column=0, columnspan=3, pady=10)

        self.count_label = tk.Label(self, textvariable=self.count_var, fg="blue")
        self.count_label.grid(row=5, column=0, columnspan=3, pady=5)

        tk.Button(self, text="Process", command=self.process_file).grid(row=4, column=1, pady=10)

    def select_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("CSV or Excel files", "*.csv *.xlsx")],
            title="Select a CSV or Excel file"
        )
        if file_path:
            ext = os.path.splitext(file_path)[1].lower()
            if ext not in [".csv", ".xlsx"]:
                messagebox.showerror("Error", "Please select a CSV (.csv) or Excel (.xlsx) file.")
                return
            if ext == ".xlsx" and openpyxl is None:
                messagebox.showerror("Error", "openpyxl is required for Excel files. Please install it with 'pip install openpyxl'.")
                return
            self.file_var.set(file_path)
            self.load_headers(file_path)
            self.reset_column_selection()
            self.update_count_label()

    def load_headers(self, file_path):
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".csv":
            with open(file_path, 'r', newline='', encoding='utf-8') as infile:
                reader = csv.reader(infile)
                headers = next(reader)
        elif ext == ".xlsx":
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        else:
            headers = []
        self.all_columns.clear()
        self.all_columns.extend(headers)
        self.column_select['values'] = self.all_columns
        self.column_select.set('')
        self.reset_column_selection()

    def reset_column_selection(self):
        for widget in self.selected_columns_frame.winfo_children():
            widget.destroy()
        self.selected_columns.clear()
        self.update_widgets.clear()
        self.update_count_label()

    def add_column(self):
        col = self.column_select.get()
        if not col or col in self.selected_columns:
            return
        self.selected_columns.append(col)
        idx = len(self.selected_columns) - 1

        row = idx
        tk.Label(self.selected_columns_frame, text=col).grid(row=row, column=0, padx=2, pady=2)
        cond_var = tk.StringVar(value="")
        cond_menu = ttk.Combobox(self.selected_columns_frame, textvariable=cond_var, width=12, state="readonly")
        cond_menu['values'] = [label for code, label in CONDITIONS]
        cond_menu.current(0)
        cond_menu.grid(row=row, column=1, padx=2, pady=2)
        filter_entry = tk.Entry(self.selected_columns_frame, width=14)
        filter_entry.grid(row=row, column=2, padx=2, pady=2)
        tk.Label(self.selected_columns_frame, text="New Value:").grid(row=row, column=3, padx=2, pady=2)
        value_entry = tk.Entry(self.selected_columns_frame, width=14)
        value_entry.grid(row=row, column=4, padx=2, pady=2)
        remove_btn = tk.Button(self.selected_columns_frame, text="Remove", command=lambda: self.remove_column(col))
        remove_btn.grid(row=row, column=5, padx=2, pady=2)
        self.update_widgets.append((col, cond_var, filter_entry, value_entry))

        cond_menu.bind("<<ComboboxSelected>>", lambda e: self.update_count_label())
        filter_entry.bind("<KeyRelease>", lambda e: self.update_count_label())
        self.update_count_label()

    def remove_column(self, col):
        if col in self.selected_columns:
            idx = self.selected_columns.index(col)
            self.selected_columns.pop(idx)
            self.update_widgets.pop(idx)
            for widget in self.selected_columns_frame.winfo_children():
                widget.destroy()
            for i, (col, cond_var, filter_entry, value_entry) in enumerate(self.update_widgets):
                tk.Label(self.selected_columns_frame, text=col).grid(row=i, column=0, padx=2, pady=2)
                cond_menu = ttk.Combobox(self.selected_columns_frame, textvariable=cond_var, width=12, state="readonly")
                cond_menu['values'] = [label for code, label in CONDITIONS]
                cond_menu.current(0)
                cond_menu.grid(row=i, column=1, padx=2, pady=2)
                filter_entry.grid(row=i, column=2, padx=2, pady=2)
                tk.Label(self.selected_columns_frame, text="New Value:").grid(row=i, column=3, padx=2, pady=2)
                value_entry.grid(row=i, column=4, padx=2, pady=2)
                remove_btn = tk.Button(self.selected_columns_frame, text="Remove", command=lambda c=col: self.remove_column(c))
                remove_btn.grid(row=i, column=5, padx=2, pady=2)
                cond_menu.bind("<<ComboboxSelected>>", lambda e: self.update_count_label())
                filter_entry.bind("<KeyRelease>", lambda e: self.update_count_label())
            self.update_count_label()

    def row_matches_filters(self, row, filters, header):
        for col, cond_var, filter_entry, _ in filters:
            if col not in header:
                return False
            idx = header.index(col)
            value = row[idx]
            cond_label = cond_var.get()
            cond_code = ""
            for code, label in CONDITIONS:
                if label == cond_label:
                    cond_code = code
                    break
            filter_value = filter_entry.get().strip()
            if not cond_code or not filter_value:
                continue
            try:
                if cond_code == "==":
                    if str(value) != filter_value:
                        return False
                elif cond_code == ">":
                    if float(value) <= float(filter_value):
                        return False
                elif cond_code == "<":
                    if float(value) >= float(filter_value):
                        return False
                elif cond_code == ">=":
                    if float(value) < float(filter_value):
                        return False
                elif cond_code == "<=":
                    if float(value) > float(filter_value):
                        return False
                elif cond_code == "contains":
                    if filter_value.lower() not in str(value).lower():
                        return False
                elif cond_code == "not contains":
                    if filter_value.lower() in str(value).lower():
                        return False
            except Exception:
                return False
        return True

    def get_updated_rows(self, input_file, updates, max_records=None):
        ext = os.path.splitext(input_file)[1].lower()
        if ext == ".csv":
            with open(input_file, 'r', newline='', encoding='utf-8') as infile:
                reader = csv.reader(infile)
                header = next(reader)
                rows = list(reader)
        elif ext == ".xlsx":
            wb = openpyxl.load_workbook(input_file, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            header = rows.pop(0)
        else:
            header = []
            rows = []

        updated_rows = []
        for row in rows:
            new_row = list(row)
            for col, cond_var, filter_entry, value_entry in updates:
                if col not in header:
                    continue
                idx = header.index(col)
                value = row[idx]
                cond_label = cond_var.get()
                cond_code = ""
                for code, label in CONDITIONS:
                    if label == cond_label:
                        cond_code = code
                        break
                filter_value = filter_entry.get().strip()
                new_value = value_entry.get().strip()
                if not cond_code or not filter_value:
                    continue
                try:
                    should_update = False
                    if cond_code == "==":
                        should_update = str(value) == filter_value
                    elif cond_code == ">":
                        should_update = float(value) > float(filter_value)
                    elif cond_code == "<":
                        should_update = float(value) < float(filter_value)
                    elif cond_code == ">=":
                        should_update = float(value) >= float(filter_value)
                    elif cond_code == "<=":
                        should_update = float(value) <= float(filter_value)
                    elif cond_code == "contains":
                        should_update = filter_value.lower() in str(value).lower()
                    elif cond_code == "not contains":
                        should_update = filter_value.lower() not in str(value).lower()
                    if should_update:
                        new_row[idx] = new_value
                except Exception:
                    continue
            updated_rows.append(new_row)

        if max_records is not None:
            updated_rows = updated_rows[:max_records]

        return header, updated_rows

    def update_count_label(self, *args):
        input_file = self.file_var.get()
        limit = self.limit_var.get().strip()
        ext = os.path.splitext(input_file)[1].lower()
        if not input_file or not os.path.isfile(input_file) or ext not in [".csv", ".xlsx"]:
            self.count_var.set("Total rows: 0")
            return
        if limit:
            try:
                max_records = int(limit)
                if max_records <= 0:
                    raise ValueError
            except ValueError:
                self.count_var.set("Total rows: 0")
                return
        else:
            max_records = None

        try:
            header, updated_rows = self.get_updated_rows(input_file, self.update_widgets, max_records)
            row_count = len(updated_rows)
        except Exception:
            row_count = 0
        self.count_var.set(f"Total rows: {row_count}")

    def process_file(self):
        input_file = self.file_var.get()
        limit = self.limit_var.get().strip()
        ext = os.path.splitext(input_file)[1].lower()
        if not input_file or not os.path.isfile(input_file) or ext not in [".csv", ".xlsx"]:
            messagebox.showerror("Error", "Please select a valid CSV (.csv) or Excel (.xlsx) file.")
            return
        if ext == ".xlsx" and openpyxl is None:
            messagebox.showerror("Error", "openpyxl is required for Excel files. Please install it with 'pip install openpyxl'.")
            return
        if limit:
            try:
                max_records = int(limit)
                if max_records <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid positive integer for the row limit, or leave blank for all.")
                return
        else:
            max_records = None

        try:
            header, updated_rows = self.get_updated_rows(input_file, self.update_widgets, max_records)
            row_count = len(updated_rows)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while updating:\n{e}")
            return

        if row_count == 0:
            messagebox.showinfo("No Rows", "No rows to process.")
            return

        confirm = messagebox.askyesno(
            "Confirm Save",
            f"{row_count} row(s) will be processed and saved.\n\nDo you want to continue?"
        )
        if not confirm:
            return

        if ext == ".csv":
            filetypes = [("CSV files", "*.csv")]
            defaultext = ".csv"
            save_func = save_csv
        else:
            filetypes = [("Excel files", "*.xlsx")]
            defaultext = ".xlsx"
            save_func = save_xlsx

        output_file = filedialog.asksaveasfilename(
            defaultextension=defaultext,
            filetypes=filetypes,
            title="Save updated file as"
        )
        if not output_file:
            return

        try:
            save_func(header, updated_rows, output_file)
            messagebox.showinfo("Success", f"Updated file saved to:\n{output_file}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")

def save_csv(header, rows, output_file):
    with open(output_file, 'w', newline='', encoding='utf-8') as outfile:
        writer = csv.writer(outfile)
        writer.writerow(header)
        writer.writerows(rows)

def save_xlsx(header, rows, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(output_file)

# --- Tkinter UI setup ---
root = tk.Tk()
root.title("CSV/Excel Editor")

# Create notebook (tabs)
notebook = ttk.Notebook(root)
notebook.pack(expand=True, fill='both', padx=5, pady=5)

# Create tabs
filter_tab = FilterTab(notebook)
update_tab = UpdateTab(notebook)

# Add tabs to notebook
notebook.add(filter_tab, text='Filter Records')
notebook.add(update_tab, text='Update Values')

root.mainloop()
